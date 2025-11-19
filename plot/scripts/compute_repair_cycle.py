#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
计算 Bug分类-FlinkCDC-ToolAdded_1119.xlsx 中 resolved bug 工作表的修复周期：
    修复周期 = 确认修复时间 - 首次报告时间
脚本假定：
    * resolved bug 表在工作簿中第二页（或已命名为 "resolved bug"）。
    * 真实表头在第二行，需要同时包含 “确认修复时间”、“首次报告时间”、“修复周期”。
运行后将就地更新 “修复周期” 列。
"""
from __future__ import annotations

from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Dict, Optional
~
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.datetime import from_excel

FILE = Path('Bug分类-FlinkCDC-ToolAdded_1119.xlsx')
TARGET_SHEET_NAME = 'resolved bug'
HEADER_ROW = 2
DATA_START_ROW = 3

COL_CONFIRM = '确认修复时间'
COL_FIRST_REPORT = '首次报告时间'
COL_CYCLE = '修复周期'


def load_sheet(path: Path) -> Worksheet:
    if not path.exists():
        raise FileNotFoundError(f'文件不存在：{path}')
    wb = openpyxl.load_workbook(path)
    if TARGET_SHEET_NAME in wb.sheetnames:
        sheet = wb[TARGET_SHEET_NAME]
    else:
        if len(wb.worksheets) < 2:
            raise RuntimeError('工作簿中缺少第二个工作表，无法定位 resolved bug 页。')
        sheet = wb.worksheets[1]
    sheet.parent.active = sheet
    return sheet


def build_header_map(sheet: Worksheet) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for col in range(1, sheet.max_column + 1):
        value = sheet.cell(row=HEADER_ROW, column=col).value
        if value is None:
            continue
        headers[str(value).strip()] = col
    return headers


def ensure_columns(headers: Dict[str, int]):
    missing = [col for col in (COL_CONFIRM, COL_FIRST_REPORT, COL_CYCLE) if col not in headers]
    if missing:
        raise RuntimeError(f'缺少必要列：{missing}')


def parse_datetime(value) -> Optional[datetime]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)):
        try:
            return from_excel(value)
        except Exception:
            return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        iso_candidate = text
        if text.upper().endswith('Z'):
            iso_candidate = text[:-1] + '+00:00'
        try:
            dt = datetime.fromisoformat(iso_candidate)
            if dt.tzinfo is not None:
                dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
            return dt
        except ValueError:
            pass
        for fmt in (
            '%Y-%m-%d %H:%M:%S',
            '%Y/%m/%d %H:%M:%S',
            '%Y-%m-%dT%H:%M:%S',
            '%Y-%m-%d',
        ):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
    return None


def format_cycle(delta: timedelta) -> str:
    total_seconds = int(delta.total_seconds())
    sign = '-' if total_seconds < 0 else ''
    total_seconds = abs(total_seconds)
    days, remainder = divmod(total_seconds, 86400)
    hours, remainder = divmod(remainder, 3600)
    minutes, seconds = divmod(remainder, 60)
    return f'{sign}{days}d {hours:02d}:{minutes:02d}:{seconds:02d}'


def compute_cycles(sheet: Worksheet, headers: Dict[str, int]):
    confirm_col = headers[COL_CONFIRM]
    first_report_col = headers[COL_FIRST_REPORT]
    cycle_col = headers[COL_CYCLE]

    updated = 0
    cleared = 0

    for row in range(DATA_START_ROW, sheet.max_row + 1):
        confirm_value = sheet.cell(row=row, column=confirm_col).value
        report_value = sheet.cell(row=row, column=first_report_col).value
        confirm_dt = parse_datetime(confirm_value)
        report_dt = parse_datetime(report_value)

        target_cell = sheet.cell(row=row, column=cycle_col)
        if confirm_dt is None or report_dt is None:
            if target_cell.value not in (None, ''):
                target_cell.value = None
                cleared += 1
            continue

        delta = confirm_dt - report_dt
        target_cell.value = format_cycle(delta)
        updated += 1

    print(f'已更新 {updated} 行修复周期，清空 {cleared} 行。')


def main():
    sheet = load_sheet(FILE)
    headers = build_header_map(sheet)
    ensure_columns(headers)
    compute_cycles(sheet, headers)
    sheet.parent.save(FILE)
    print('✅ 修复周期计算完成。')


if __name__ == '__main__':
    main()
